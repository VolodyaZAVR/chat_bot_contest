import os
import sqlite3
import openpyxl

'''
def export_to_sqlite():
    #Экспорт данных из xlsx в sqlite

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'code_station.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS codes (id text, station text, region text)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('station_code.xlsx', data_only=True)
    sheet = file_to_read['Лист1']
    print(sheet.max_row)
    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 3 ( 4 не включая)
        for col in range(1, 4):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO codes VALUES (?, ?, ?);", (data[0], data[1], data[2]))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    #Очистка базы sqlite

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'code_station.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM codes")
    connect.commit()
    connect.close()
'''
''' bd3
def export_to_sqlite():
    #Экспорт данных из xlsx в sqlite

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'filials.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS filials (name text, contatcts text, address text, region text, station text, carriage text)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('filials.xlsx', data_only=True)
    sheet = file_to_read['Лист1']
    print(sheet.max_row)
    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 7):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO filials VALUES (?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4], data[5]))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    #Очистка базы sqlite

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'filials.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM filials")
    connect.commit()
    connect.close()
'''


# bd 2 region - station
def export_to_sqlite():
    # Экспорт данных из xlsx в sqlite
    # 1. Создание и подключение к базе
    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    base_name = 'stations.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS stations (region text, station text, region_lower text)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('region.xlsx', data_only=True)
    sheet = file_to_read['Лист1']
    print(sheet.max_row)
    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 4):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO stations VALUES (?, ?, ?);", (data[0], data[1], data[2].lower()))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    # Очистка базы sqlite

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'stations.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM stations")
    connect.commit()
    connect.close()


clear_base()
# Запуск функции
export_to_sqlite()
